# -*- coding: utf-8 -*-
"""
Arma productos.json para el buscador de precios Pewen a partir de la lista
minorista (xlsx) + el indice de fotos ya armado en buscador-fotos-pewen/fotos.db.

Uso:
    python build_catalogo.py "ruta a la Lista Minorista.xlsx"

Vuelve a correr esto cada vez que haya productos nuevos: pisa productos.json,
SIN-FOTO.csv y AGRUPADOS.csv.
"""
import sys
import re
import json
import csv
import sqlite3
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
FOTOS_PEWEN_DIR = HERE.parent / "buscador-fotos-pewen"
FOTOS_DB = FOTOS_PEWEN_DIR / "fotos.db"
FOTOS_OUT_DIR = HERE / "fotos"

SHEET_NAME = "Lista Definitiva (365)"

# Columnas (0-based) de la hoja, confirmadas contra el encabezado de fila 2.
COL_CODIGO, COL_NOMBRE, COL_RUBRO, COL_LINEA, COL_MEDIDAS, COL_BASE, COL_PRECIO, COL_MONEDA = range(8)
COL_STOCK = 37

TIER_LABELS = ["0-50 m2", "50-100 m2", "100-200 m2", "200-250 m2", "250+ (mayorista)"]
PAGO_LABELS = ["contado", "tarjeta1", "tarjeta3", "tarjeta6", "tarjeta12"]
PAGO_COLS = {
    "contado": range(11, 16),
    "tarjeta1": range(16, 21),
    "tarjeta3": range(21, 26),
    "tarjeta6": range(26, 31),
    "tarjeta12": range(31, 36),
}

EXCLUDE_RUBROS = {
    "Zocalos MDF", "Zocalos Vinilico", "Terminaciones", "ACCESORIOS Reve Pared",
    "ACCESORIOS DECK", "Weber", "adhesivos", "Lacas y Adhesivos", "Mantos",
    "Perfiles WPC", "Muebles Jardin Exterior", "masa niveladora", "Pegamento",
    "Accesorios Pared PVC", "FIELTRO", "prymer", "Recubrimiento Superficies rusticas ",
    "Bajo Piso", "Servicios", "ACOPIO", "cesped", "Accesorios Quick-Step",
    "AUTOPOSANTE", "Limipiador SPC", "IMPRIMACION", "autonivelante",
    "Revestimieto de Pileta",
}

# (rubro, linea) puntuales que hay que sacar del catalogo aunque su rubro en
# general si entre: son accesorios/terminaciones sueltos dentro de un rubro
# que por lo demas es piso de verdad (ej. "Baldosa Encastrable Deck WPC" trae
# baldosas reales + los angulos/esquineros de esa misma baldosa).
EXCLUDE_LINEAS = {
    ("Baldosa Encastrable Deck WPC", "Angulos"),
    ("homogeneos", "zocalos homogeneos"),
    ("homogeneos", "accesorios homogeneos"),
}

# (rubro, linea) puntuales cuyo Rubro en la planilla es enganoso: comparten
# rubro con un piso pero en realidad son revestimiento de pared (Slimstone/
# Slimwood son placas decorativas de pared que quedaron cargadas como
# "Placa SPC"; "pared homogeneo/heterogenea" son vinilico de PARED sueltos
# dentro del rubro "homogeneos" que es mayormente piso).
LINEA_GROUP_OVERRIDE = {
    ("Placa SPC", "Slimstone"): "Revestimiento de pared",
    ("Placa SPC", "Slimwood"): "Revestimiento de pared",
    ("homogeneos", "pared heterogenea"): "Revestimiento de pared",
    ("homogeneos", "pared homogeneo"): "Revestimiento de pared",
}

# Rubro -> grupo/categoria amigable para los chips del buscador.
GROUP_MAP = [
    (("Flotante", "ORCA"), "Flotante"),
    (("SPC", "Placa SPC"), "SPC"),
    (("LVT",), "LVT / Vinílico click"),
    (("Porcelanato",), "Porcelanato"),
    (("Piso Ingenieria", "Prefinish", "Undfinish", "Tablones", "Deck de Madera"), "Piso de madera"),
    (("Deck", "WPC", "TAPA DECK", "Baldosa Encastrable"), "Deck WPC"),
    (("Revestimiento Pared",), "Revestimiento de pared"),
    (("Ceramico",), "Cerámico"),
    (("Vinilico", "VINILICO", "homogeneos", "rollo"), "Vinílico rollo/pegar"),
    (("DECO",), "Deco"),
]


def grupo_de(rubro, linea):
    override = LINEA_GROUP_OVERRIDE.get((rubro, linea))
    if override:
        return override
    for prefixes, grupo in GROUP_MAP:
        if any(rubro.startswith(p) or p in rubro for p in prefixes):
            return grupo
    return "Otros"


# Rubro (o su prefijo) -> carpetas de Drive candidatas para buscar foto.
FOLDER_MAP = [
    (("Flotante", "ORCA"), ("Flotante", "0.PREVENTA", "escaneado", "nuevos tonos stock")),
    (("SPC", "Placa SPC"), ("SPC", "outlet", "nuevos tonos stock")),
    (("LVT",), ("LVT", "nueva sesion de fotos", "fotos guada vinilicos", "outlet")),
    (("Porcelanato",), ("Porcelanatos",)),
    (("Deck", "WPC", "TAPA DECK", "Baldosa Encastrable"), ("Decks",)),
    (("Revestimiento Pared",), ("Revestimiento Pared", "revs", "silenza")),
    (("Piso Ingenieria", "Prefinish", "Undfinish", "Tablones"), ("FOTOS DE PRODUCTOS", "catalogos")),
    (("Ceramico",), ("FOTOS DE PRODUCTOS",)),
    (("Vinilico", "VINILICO", "homogeneos", "rollo"), ("fotos guada vinilicos", "FOTOS DE PRODUCTOS")),
    (("DECO",), ("FOTOS DE PRODUCTOS", "fotos para publi")),
]

LINEA_FOLDER_OVERRIDE = {
    ("Placa SPC", "Slimstone"): ("Revestimiento Pared", "revs", "silenza"),
    ("Placa SPC", "Slimwood"): ("Revestimiento Pared", "revs", "silenza"),
    ("homogeneos", "pared heterogenea"): ("Revestimiento Pared", "revs", "silenza"),
    ("homogeneos", "pared homogeneo"): ("Revestimiento Pared", "revs", "silenza"),
}

STOPWORDS = {
    "CON", "SIN", "PARA", "DEL", "LOS", "LAS", "UNA", "PISO", "PISOS", "OFERTA",
    "LINEA", "PRIMERA", "SEGUNDA", "DECK", "DECKS", "WPC", "REVESTIMIENTO",
    "PARED", "PORCELANATO", "FLOTANTE", "VINILICO", "COLECCION", "PLACA",
    "TABLA", "TABLAS", "TABLON", "TABLONES", "EXTERIOR", "INTERIOR", "CERAMICO",
}


def normalize_group_name(nombre):
    s = (nombre or "").upper()
    s = re.sub(r"\([^)]*\d[^)]*\)", " ", s)
    s = re.sub(r"\d+[.,]?\d*\s*[X\u00d7*]\s*\d+[.,]?\d*(\s*[X\u00d7*]\s*\d+[.,]?\d*)?\s*(MM|CM|M2|M\u00b2|ML|M)?\b", " ", s)
    s = re.sub(r"\bX\s*\d+[.,]?\d*\s*(MM|CM|M2|M\u00b2|ML|M)?\b", " ", s)
    s = re.sub(r"\d+[.,]?\d*\s*(MM|CM|M2|M\u00b2|ML)\b", " ", s)
    s = re.sub(r"[-\u2013]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def moneda_norm(m):
    if not m:
        return "ARS"
    m = str(m).strip().upper()
    if m.startswith("USD"):
        return "USD"
    if m == "EURO":
        return "EUR"
    return "ARS"


def unidad_de(rubro):
    rubro_l = (rubro or "").lower()
    if any(k in rubro_l for k in ("zocalo", "terminacion", "tapajunta", "angulo")):
        return "ml"
    return "m2"


DIM_RE = re.compile(r"^\d+X\d+$")


def significant_words(nombre):
    s = re.sub(r"[^A-Z0-9\u00c1\u00c9\u00cd\u00d3\u00da\u00d1 ]", " ", (nombre or "").upper())
    words = [w for w in s.split() if len(w) >= 4 and w not in STOPWORDS and not DIM_RE.match(w)]
    return words


def num_or_none(v):
    if isinstance(v, (int, float)):
        return round(v, 2)
    return None


def codigo_str(v):
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def carpeta_top(path):
    return path.split("\\")[0].split("/")[0]


def cargar_fotos():
    con = sqlite3.connect(str(FOTOS_DB))
    cur = con.cursor()
    cur.execute("SELECT id, name, url_view, thumb_path FROM fotos WHERE hidden=0")
    fotos = []
    for fid, name, url_view, thumb_path in cur.fetchall():
        fotos.append({"id": fid, "name": name or "", "carpeta": carpeta_top(name or ""), "url": url_view, "thumb": thumb_path})
    con.close()
    return fotos


def copiar_thumb(foto):
    """Copia el thumbnail ya bajado por buscador-fotos-pewen a fotos/ local. Devuelve la ruta relativa o None."""
    if not foto.get("thumb"):
        return None
    origen = FOTOS_PEWEN_DIR / foto["thumb"]
    if not origen.exists():
        return None
    FOTOS_OUT_DIR.mkdir(exist_ok=True)
    destino = FOTOS_OUT_DIR / (foto["id"] + origen.suffix)
    if not destino.exists():
        shutil.copyfile(origen, destino)
    return "fotos/" + destino.name


def folders_para_rubro(rubro, linea):
    override = LINEA_FOLDER_OVERRIDE.get((rubro, linea))
    if override:
        return override
    for prefixes, folders in FOLDER_MAP:
        if any(rubro.startswith(p) or p in rubro for p in prefixes):
            return folders
    return ()


def buscar_foto(rubro, linea, nombre, fotos_por_carpeta):
    folders = folders_para_rubro(rubro, linea)
    if not folders:
        return None
    linea_words = set(significant_words(linea))
    words = set(significant_words(nombre)) - linea_words
    if not words:
        return None
    mejor = None
    mejor_score = 0
    mejor_con_codigo = False
    for folder in folders:
        for foto in fotos_por_carpeta.get(folder, []):
            fwords = set(significant_words(foto["name"].replace("\\", " ").replace("/", " "))) - linea_words
            overlap = words & fwords
            con_codigo = any(any(c.isdigit() for c in w) for w in overlap)
            score = sum(2 if any(c.isdigit() for c in w) else 1 for w in overlap)
            if score > mejor_score:
                mejor_score = score
                mejor = foto
                mejor_con_codigo = con_codigo
    # con codigo numerico de por medio, 3 alcanza (el codigo ya es bastante especifico);
    # sin codigo, pedimos mas superposicion de palabras para evitar falsos positivos
    # por palabras genericas de linea/coleccion.
    umbral = 3 if mejor_con_codigo else 4
    if mejor_score >= umbral:
        return mejor
    return None


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not xlsx_path:
        print("Uso: python build_catalogo.py \"ruta al xlsx\"")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    filas = []
    excluidas_rubro = 0
    excluidas_linea = 0
    excluidas_sin_precio = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[COL_CODIGO] is None:
            continue
        rubro = (row[COL_RUBRO] or "").strip()
        linea = (row[COL_LINEA] or "").strip()
        if rubro in EXCLUDE_RUBROS:
            excluidas_rubro += 1
            continue
        if (rubro, linea) in EXCLUDE_LINEAS:
            excluidas_linea += 1
            continue
        if not row[COL_PRECIO]:
            excluidas_sin_precio += 1
            continue
        filas.append(row)

    print(f"Filas totales con codigo: {excluidas_rubro + excluidas_linea + excluidas_sin_precio + len(filas)}")
    print(f"  excluidas por rubro (zocalos/accesorios/etc): {excluidas_rubro}")
    print(f"  excluidas por linea puntual (accesorios sueltos): {excluidas_linea}")
    print(f"  excluidas sin precio de lista: {excluidas_sin_precio}")
    print(f"  quedan: {len(filas)}")

    grupos = defaultdict(list)
    for row in filas:
        rubro = (row[COL_RUBRO] or "").strip()
        linea = (row[COL_LINEA] or "").strip()
        key = (rubro, linea, normalize_group_name(row[COL_NOMBRE]))
        grupos[key].append(row)

    print(f"Productos agrupados (cards finales): {len(grupos)}")
    multi = {k: v for k, v in grupos.items() if len(v) > 1}
    print(f"  de los cuales con 2+ medidas: {len(multi)}")

    with open(HERE / "AGRUPADOS.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["rubro", "linea", "nombre_base", "codigo", "nombre_original", "medidas", "precio_lista", "moneda"])
        for key, rows in sorted(multi.items()):
            for r in rows:
                w.writerow([key[0], key[1], key[2], codigo_str(r[COL_CODIGO]), r[COL_NOMBRE], r[COL_MEDIDAS] or "",
                            num_or_none(r[COL_PRECIO]), moneda_norm(r[COL_MONEDA])])

    fotos = cargar_fotos()
    fotos_por_carpeta = defaultdict(list)
    for foto in fotos:
        fotos_por_carpeta[foto["carpeta"]].append(foto)
    print(f"Fotos indexadas disponibles (fotos.db): {len(fotos)}")

    productos = []
    con_foto = 0
    sin_foto_rows = []

    for (rubro, linea, nombre_base), rows in sorted(grupos.items()):
        rows_sorted = sorted(rows, key=lambda r: (r[COL_MEDIDAS] or ""))
        primero = rows_sorted[0]
        moneda = moneda_norm(primero[COL_MONEDA])

        variantes = []
        for r in rows_sorted:
            precios = {}
            for pago, cols in PAGO_COLS.items():
                precios[pago] = [num_or_none(r[c]) for c in cols]
            variantes.append({
                "codigo": codigo_str(r[COL_CODIGO]),
                "nombre": r[COL_NOMBRE],
                "medidas": (r[COL_MEDIDAS] or "").strip() or None,
                # "Precio de Lista" (columna F) es el precio sin descuento, igual al de
                # tarjeta en 12 pagos. Lo que se cobra de verdad es la columna "0-50 m2"
                # de contado/transferencia (ya trae el descuento admisible aplicado).
                "precio_venta": precios["contado"][0],
                "precio_lista_bruto": num_or_none(r[COL_PRECIO]),
                "moneda": moneda_norm(r[COL_MONEDA]),
                "stock": r[COL_STOCK] if isinstance(r[COL_STOCK], (int, float)) else None,
                "precios": precios,
            })

        foto = buscar_foto(rubro, linea, primero[COL_NOMBRE], fotos_por_carpeta)
        foto_local = copiar_thumb(foto) if foto else None
        if foto_local:
            con_foto += 1
        else:
            sin_foto_rows.append([rubro, linea, nombre_base, primero[COL_NOMBRE]])

        productos.append({
            "id": codigo_str(primero[COL_CODIGO]),
            "nombre_base": nombre_base.title(),
            "rubro": rubro,
            "grupo": grupo_de(rubro, linea),
            "linea": linea,
            "unidad": unidad_de(rubro),
            "moneda": moneda,
            "precio_desde": min((v["precio_venta"] for v in variantes if v["precio_venta"] is not None), default=None),
            "stock_total": sum(v["stock"] for v in variantes if v["stock"] is not None) or None,
            "foto_url": foto_local,
            "foto_drive_url": foto["url"] if foto else None,
            "variantes": variantes,
        })

    with open(HERE / "productos.json", "w", encoding="utf-8") as f:
        json.dump({
            "generado_de": Path(xlsx_path).name,
            "tier_labels": TIER_LABELS,
            "pago_labels": PAGO_LABELS,
            "productos": productos,
        }, f, ensure_ascii=False, separators=(",", ":"))

    with open(HERE / "SIN-FOTO.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["rubro", "linea", "nombre_base", "nombre_ejemplo"])
        w.writerows(sin_foto_rows)

    print(f"Con foto: {con_foto}/{len(productos)} ({con_foto*100//len(productos)}%)")
    print(f"Sin foto: {len(sin_foto_rows)} -> SIN-FOTO.csv")
    print("Escrito productos.json, AGRUPADOS.csv, SIN-FOTO.csv")


if __name__ == "__main__":
    main()
